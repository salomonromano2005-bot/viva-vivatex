"""
Generador de Excel profesional para AVIVA - Grupo Vivatex S.A. de C.V.
Uso: python3 generate_excel.py <json_base64> <output_path>
"""
import sys
import json
import base64
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.utils import get_column_letter
import datetime

VERDE      = "1E6B2E"
VERDE_MED  = "4A9A20"
VERDE_L    = "D4EDDA"
VERDE_XL   = "F0F9F0"
BLANCO     = "FFFFFF"
NEGRO      = "1C1C1C"
GRIS_OSC   = "4A4A4A"
GRIS_L     = "F5F5F5"
AZUL_H     = "1A3A5C"

def fill(c): return PatternFill("solid", fgColor=c)
def border():
    s = Side(style='thin', color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)
def font(size=10, bold=False, color=NEGRO, italic=False):
    return Font(name='Calibri', size=size, bold=bold, color=color, italic=italic)
def align(h='left', v='center', wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def portada(wb, titulo, usuario, periodo, subtitulo):
    ws = wb.create_sheet("Portada", 0)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 55
    ws.column_dimensions['C'].width = 25

    # Fondo verde encabezado
    for r in range(1, 5):
        for col in ['A','B','C','D']:
            ws[f'{col}{r}'].fill = fill(VERDE)

    ws.row_dimensions[2].height = 55
    ws.row_dimensions[3].height = 35
    ws.merge_cells('B2:C2')
    ws['B2'].value = "GRUPO VIVATEX S.A. DE C.V."
    ws['B2'].font = Font(name='Calibri', size=22, bold=True, color=BLANCO)
    ws['B2'].alignment = align('left')

    ws.merge_cells('B3:C3')
    ws['B3'].value = "AVIVA — Sistema de Inteligencia Empresarial"
    ws['B3'].font = Font(name='Calibri', size=12, color="AADDAA", italic=True)
    ws['B3'].alignment = align('left')

    # Línea divisora
    for col in ['A','B','C','D']:
        ws[f'{col}5'].fill = fill(VERDE_MED)
        ws.row_dimensions[5].height = 5

    ws.row_dimensions[7].height = 35
    ws.merge_cells('B7:C7')
    ws['B7'].value = titulo.upper()
    ws['B7'].font = Font(name='Calibri', size=18, bold=True, color=VERDE)
    ws['B7'].alignment = align('left')

    if subtitulo:
        ws.row_dimensions[8].height = 22
        ws.merge_cells('B8:C8')
        ws['B8'].value = subtitulo
        ws['B8'].font = Font(name='Calibri', size=11, color=GRIS_OSC, italic=True)
        ws['B8'].alignment = align('left')

    meta = [
        ("Período:", periodo),
        ("Generado por:", "AVIVA · Inteligencia Artificial Vivatex"),
        ("Fecha:", datetime.datetime.now().strftime("%d/%m/%Y %H:%M")),
        ("Usuario:", usuario),
        ("Confidencial:", "Uso exclusivo interno — Grupo Vivatex S.A. de C.V."),
    ]
    for i, (label, val) in enumerate(meta, 10):
        ws.row_dimensions[i].height = 22
        ws[f'B{i}'].value = label
        ws[f'B{i}'].font = Font(name='Calibri', size=10, bold=True, color=GRIS_OSC)
        ws[f'C{i}'].value = val
        ws[f'C{i}'].font = Font(name='Calibri', size=10, color=NEGRO)

def hoja_datos(wb, hoja_data, titulo, usuario, periodo):
    nombre = str(hoja_data.get("nombre", "Datos"))[:31]
    columnas = hoja_data.get("columnas", [])
    filas = hoja_data.get("filas", [])
    totales = hoja_data.get("totales", None)
    tipo_graf = hoja_data.get("tipo_grafica", None)

    ws = wb.create_sheet(nombre)
    ws.sheet_view.showGridLines = False

    n_cols = max(len(columnas), 1)

    # Encabezado empresa
    ws.row_dimensions[1].height = 10
    ws.row_dimensions[2].height = 45
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
    c = ws.cell(row=2, column=1)
    c.value = f"GRUPO VIVATEX S.A. DE C.V.  ·  {titulo.upper()}"
    c.font = Font(name='Calibri', size=13, bold=True, color=BLANCO)
    c.fill = fill(VERDE)
    c.alignment = align('center')

    ws.row_dimensions[3].height = 5
    ws.row_dimensions[4].height = 18
    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=n_cols)
    c = ws.cell(row=4, column=1)
    c.value = f"{nombre}   |   {periodo}   |   Generado {datetime.datetime.now().strftime('%d/%m/%Y %H:%M')}   |   Usuario: {usuario}"
    c.font = Font(name='Calibri', size=9, italic=True, color=GRIS_OSC)
    c.fill = fill(VERDE_L)
    c.alignment = align('center')

    # Anchos de columna automáticos
    for j, col_name in enumerate(columnas, 1):
        max_w = max(len(str(col_name)) + 4, 12)
        for fila in filas:
            if j-1 < len(fila):
                max_w = max(max_w, min(len(str(fila[j-1])) + 2, 42))
        ws.column_dimensions[get_column_letter(j)].width = max_w

    # Headers columnas
    HR = 6
    ws.row_dimensions[HR].height = 28
    for j, h in enumerate(columnas, 1):
        c = ws.cell(row=HR, column=j)
        c.value = h
        c.font = Font(name='Calibri', size=11, bold=True, color=BLANCO)
        c.fill = fill(AZUL_H)
        c.alignment = align('center')
        c.border = border()

    # Detectar columnas numéricas
    num_cols = set()
    for j in range(1, len(columnas)+1):
        if any(isinstance(f[j-1], (int, float)) for f in filas if j-1 < len(f)):
            num_cols.add(j)

    # Filas de datos
    DS = HR + 1
    for i, fila in enumerate(filas):
        r = DS + i
        ws.row_dimensions[r].height = 20
        bg = VERDE_XL if i % 2 == 1 else BLANCO
        for j, val in enumerate(fila, 1):
            c = ws.cell(row=r, column=j)
            c.value = val
            c.font = Font(name='Calibri', size=10)
            c.fill = fill(bg)
            c.border = border()
            if isinstance(val, (int, float)):
                c.alignment = align('right', wrap=False)
                if j in num_cols:
                    if isinstance(val, float) and abs(val) < 1 and val != 0:
                        c.number_format = '0.00%'
                    else:
                        c.number_format = '#,##0.00'
            else:
                c.alignment = align('left')

    # Fila totales
    if totales:
        tr = DS + len(filas) + 1
        ws.row_dimensions[tr].height = 24
        for j, val in enumerate(totales, 1):
            c = ws.cell(row=tr, column=j)
            c.value = val
            c.font = Font(name='Calibri', size=11, bold=True, color=BLANCO)
            c.fill = fill(VERDE_MED)
            c.alignment = align('center')
            c.border = border()

    # Gráfica
    if tipo_graf and len(filas) >= 2 and len(columnas) >= 2:
        try:
            val_cols = [j for j in range(2, min(len(columnas)+1, 6))
                        if any(isinstance(f[j-1], (int, float)) for f in filas if j-1 < len(f))]
            if not val_cols:
                return

            if tipo_graf == "bar":
                chart = BarChart()
                chart.type = "col"
                chart.grouping = "clustered"
                chart.style = 10
            elif tipo_graf == "line":
                chart = LineChart()
                chart.style = 10
            elif tipo_graf == "pie":
                chart = PieChart()
                chart.style = 10
            else:
                chart = BarChart()
                chart.type = "col"
                chart.style = 10

            chart.title = nombre
            chart.width = 22
            chart.height = 14

            if tipo_graf in ["bar", "line"]:
                for vc in val_cols[:3]:
                    dr = Reference(ws, min_col=vc, min_row=HR, max_col=vc, max_row=HR+len(filas))
                    chart.add_data(dr, titles_from_data=True)
                cats = Reference(ws, min_col=1, min_row=DS, max_row=DS+len(filas)-1)
                chart.set_categories(cats)
                if hasattr(chart, 'y_axis'):
                    chart.y_axis.title = columnas[val_cols[0]-1] if val_cols else ""
                    chart.x_axis.title = columnas[0]
            elif tipo_graf == "pie" and val_cols:
                dr = Reference(ws, min_col=val_cols[0], min_row=HR, max_col=val_cols[0], max_row=HR+len(filas))
                chart.add_data(dr, titles_from_data=True)
                cats = Reference(ws, min_col=1, min_row=DS, max_row=DS+len(filas)-1)
                chart.set_categories(cats)

            chart_row = DS + len(filas) + (3 if not totales else 4)
            ws.add_chart(chart, f"A{chart_row}")
        except Exception as e:
            pass  # Si la gráfica falla, continuar sin ella

def generate(data):
    wb = Workbook()
    wb.remove(wb.active)

    titulo    = data.get("titulo", "Reporte Vivatex")
    usuario   = data.get("usuario", "Usuario")
    periodo   = data.get("periodo", datetime.datetime.now().strftime("%B %Y"))
    subtitulo = data.get("subtitulo", "")

    portada(wb, titulo, usuario, periodo, subtitulo)

    for hoja in data.get("hojas", []):
        hoja_datos(wb, hoja, titulo, usuario, periodo)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()

if __name__ == "__main__":
    try:
        # El servidor manda los datos por stdin como JSON
        raw = sys.stdin.read().strip()
        data = json.loads(raw)
        
        # output_path puede venir como 2do argumento
        output_path = sys.argv[1] if len(sys.argv) > 1 else None
        
        result = generate(data)
        
        if output_path:
            with open(output_path, 'wb') as f:
                f.write(result)
            print(json.dumps({"success": True, "path": output_path}))
        else:
            print(base64.b64encode(result).decode())
    except Exception as e:
        print(json.dumps({"success": False, "error": str(e)}))
        sys.exit(1)
